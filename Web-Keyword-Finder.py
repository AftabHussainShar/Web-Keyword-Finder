import openpyxl
import requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse, urljoin


keywords = ["nesten", "nest planning", "planning zomer","plannen","nestplannen","planning"]

# def is_media_url(url):
#     media_extensions = {'.jpg', '.jpeg','.pdf', '.png', '.gif', '.bmp', '.svg', '.mp3', '.mp4', '.avi', '.mov'}
#     return any(url.lower().endswith(ext) for ext in media_extensions)

# def is_media_url(url):
def is_media_url(url):
    media_extensions = {'.jpg', '.jpeg', '.pdf', '.png', '.gif', '.bmp', '.svg', '.mp3', '.mp4', '.avi', '.mov'}
    return any(ext in url.lower() for ext in media_extensions)

# def check_keywords(url, keywords):
def check_keywords(url, keywords):
    try:
        response = requests.get(url)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')

        result = "No"
        for keyword in keywords:
            if keyword.lower() in soup.get_text().lower():
                result = "Yes"
                break

        return result

    except requests.exceptions.RequestException as e:
        return "Not accessible"

def process_url(url, keywords, row_index, sheet, visited_urls, master_domain):
    if urlparse(url).scheme not in ['http', 'https'] or url.lower() == 'javascript:void(0);':
        return

    if url in visited_urls:
        return

    if url == 'https://www.hetriethofje.nl/':
        return

    visited_urls.add(url)
    if master_domain not in urlparse(url).netloc:
        return

    if master_domain != urlparse(url).netloc:
        return

    if is_media_url(url):
        return

    result = check_keywords(url, keywords)
    sheet.cell(row=row_index, column=sheet.max_column, value=result)
    print(f"URL: {url} Result: {result}")

    if result == "Yes":
        return
    try:
        response = requests.get(url)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')

        sub_links = [urljoin(url, a['href']) for a in soup.find_all('a', href=True)]

        for sub_link in sub_links:
            process_url(sub_link, keywords, row_index, sheet, visited_urls, master_domain)
            if sheet.cell(row=row_index, column=sheet.max_column).value == "Yes":
                break

    except requests.exceptions.RequestException as e:
        print(f"Error accessing sub-links for {url}: {e}")

workbook = openpyxl.load_workbook('websites.xlsx')
sheet = workbook.active
urls = [cell.value for cell in sheet['A'][1:]]
urls = [url for url in urls if url is not None]

result_column_letter = 'E'

if result_column_letter not in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=1, max_row=1):
    sheet.insert_cols(sheet.max_column + 1)
    sheet.cell(row=1, column=sheet.max_column, value='Nesten')

if urls:
    for url in urls:
        row_index = urls.index(url) + 2
        visited_urls = set()
        master_domain = urlparse(url).netloc
        process_url(url, keywords, row_index, sheet, visited_urls, master_domain)

workbook.save('websites_with_results.xlsx')
